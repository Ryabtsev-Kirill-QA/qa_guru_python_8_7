import os
from PyPDF2 import PdfReader
from xlrd import open_workbook
from openpyxl import load_workbook
import zipfile
from utils import TMP_PATH, RESOURCES_PATH


def test_names_of_files():
    files_in_dir = os.listdir(RESOURCES_PATH)

    with zipfile.ZipFile("tmp/test.zip") as zip_file:
        assert files_in_dir == zip_file.namelist()


def test_txt_in_zip():
    text_size = os.path.getsize(os.path.join(RESOURCES_PATH, "hello.txt"))

    with open(os.path.join(RESOURCES_PATH, "hello.txt"), encoding='latin-1') as f:
        txt_file_text = f.read()

    with zipfile.ZipFile(os.path.join(TMP_PATH, 'test.zip')) as txt_z:
        assert text_size == txt_z.getinfo('hello.txt').file_size
        assert txt_file_text in txt_z.read('hello.txt').decode('latin-1')


def test_pdf_file_in_zip():
    reader = PdfReader(os.path.join(RESOURCES_PATH, "Python Testing with Pytest (Brian Okken).pdf"))
    size_of_pdf = os.path.getsize(os.path.join(RESOURCES_PATH, "Python Testing with Pytest (Brian Okken).pdf"))
    number_of_pages = len(reader.pages)
    first_page = reader.pages[1]
    last_page = reader.pages[240]
    text_first_page = first_page.extract_text()
    text_last_page = last_page.extract_text()

    with zipfile.ZipFile(os.path.join(TMP_PATH, 'test.zip')) as pdf_z:
        assert size_of_pdf == pdf_z.getinfo("Python Testing with Pytest (Brian Okken).pdf").file_size
        assert number_of_pages == len(PdfReader(pdf_z.open("Python Testing with Pytest (Brian Okken).pdf")).pages)
        assert text_first_page == PdfReader(pdf_z.open("Python Testing with Pytest (Brian Okken).pdf")).pages[1].extract_text()
        assert text_last_page == PdfReader(pdf_z.open("Python Testing with Pytest (Brian Okken).pdf")).pages[240].extract_text()


def test_xls_in_zip():
    book = open_workbook(os.path.join(RESOURCES_PATH, "file_example_XLS_10.xls"))
    size_of_xls = os.path.getsize(os.path.join(RESOURCES_PATH, "file_example_XLS_10.xls"))
    number_of_sheets_xls = book.nsheets
    names_of_sheets_xls = book.sheet_names()
    sheet_xls = book.sheet_by_index(0)
    number_of_rows_xls = sheet_xls.nrows
    number_of_rows_cols_xls = sheet_xls.ncols
    sheet_value_xls = sheet_xls.cell_value(3, 3)

    with zipfile.ZipFile(os.path.join(TMP_PATH, 'test.zip')) as xls_z:
        assert size_of_xls == xls_z.getinfo("file_example_XLS_10.xls").file_size
        assert number_of_sheets_xls == open_workbook(file_contents=xls_z.read("file_example_XLS_10.xls")).nsheets
        assert names_of_sheets_xls == open_workbook(file_contents=xls_z.read("file_example_XLS_10.xls")).sheet_names()
        assert number_of_rows_xls == open_workbook(file_contents=xls_z.read("file_example_XLS_10.xls")).sheet_by_index(0).nrows
        assert number_of_rows_cols_xls == open_workbook(file_contents=xls_z.read("file_example_XLS_10.xls")).sheet_by_index(0).ncols
        assert sheet_value_xls == open_workbook(file_contents=xls_z.read("file_example_XLS_10.xls")).sheet_by_index(0).cell_value(3, 3)


def test_xlsx_file_in_zip():
    workbook = load_workbook(os.path.join(RESOURCES_PATH, "file_example_XLSX_50.xlsx"))
    size_of_xlsx = os.path.getsize(os.path.join(RESOURCES_PATH, "file_example_XLSX_50.xlsx"))
    number_of_sheets_xlsx = len(workbook.sheetnames)
    names_of_sheets_xlsx = workbook.sheetnames
    sheet = workbook.active
    sheet_value_xlsx = sheet.cell(row=3, column=3).value

    with zipfile.ZipFile(os.path.join(TMP_PATH, 'test.zip')) as xlsx_z:
        assert size_of_xlsx == xlsx_z.getinfo("file_example_XLSX_50.xlsx").file_size
        assert number_of_sheets_xlsx == len(load_workbook(xlsx_z.open("file_example_XLSX_50.xlsx")).sheetnames)
        assert names_of_sheets_xlsx == load_workbook(xlsx_z.open("file_example_XLSX_50.xlsx")).sheetnames
        assert sheet_value_xlsx == load_workbook(xlsx_z.open("file_example_XLSX_50.xlsx")).active.cell(row=3, column=3).value
